# I am unable to make multiple requests at once and update them becuase the free API service doesn't provide this faciity. Subscription will be required for that. 
'''
Error Message Recieved while making bulk queries.
{
  "success":false,"error":{
    "code":604,
    "type":"bulk_queries_not_supported_on_plan",
    "info":"Your current subscription plan does not support bulk queries. Please upgrade your account to use this feature."}}
'''

import requests
from pathlib import Path
import openpyxl, time , _thread

def toFah(cel):
  '''
  Function to Convert the temerature from Celsius to Fahrenheit.
  '''
  return (cel*9//5)+32

def terminator(flag):
  '''
  A thread that will run in parallel. Used for safe Termination of program.
  It "takes an empty list as an Argument" and appends "True" to it whenever "Enter" is pressed.

  '''
  print("Welcome!!")
  print("System is starting. If anytime you wish to exit just press ENTER. [WARNING: Force Closing may lead to currupting the excel file.]")
  print("---------------------------------------------------------------------------------------------------------------------------------")
  input()
  flag.append(False)
  print("The System will terminate shortly. Thank You!!")

def get_Value(loc):
  '''
  Method to generate requests to API , recieve the response and extract needed data from it and return.
  It takes name of city as a string Argument.
  '''

  __params = {
  'access_key': 'efa59c8305d828986c91c19480a592a5',
  'query': loc
  }

  api_result = requests.get('http://api.weatherstack.com/current', __params)

  api_response = api_result.json()

  # print(api_response)

  return [api_response['request']['query'], api_response['current']['humidity'], api_response['current']['temperature']]


def main():
  '''
  This is the main function. Takes no arguments.
  
  Functions of this method:
  - Reads the xlsx file for useful Names of Cities, Update Status, Units (Celsius of Fahrenheit).
  - Launches terminator Thread. 
  - Calls the get_Value(loc) method for recieving the the Temp, Humidity.
  - Converts the temperature into required unit.
  - Saves the data into xlsx file.

  '''

  xlsx_file = Path('D:\Programming\Python\Projects\Live Weather Monitor\Values.xlsx')
  wb_obj = openpyxl.load_workbook(xlsx_file) 
  sheet = wb_obj.active

  # For initial update. When True all the rows will be updated, whether or not "Update(0/1)" column is 1 or 0.
  initial = True

  # To store Total Updated rows in last attempt. Will be used to terminate the program, In case no rows are being updated.
  tupdated = 0

  # Number of Rows sheet.
  n = sheet.max_row

  # Empty List used for terminator thread.
  flag =[]

  # Thread Launch
  _thread.start_new_thread(terminator, (flag,))

  # Run until the flag list is Empty.
  while not flag:
    tupdated = 0

    # Run for Each Row
    for i in range(1, n):

      # if flag is empty and either this is an initial update or "Update(0/1)" column in sheet is 1
      if ((sheet.cell(row = i+1, column = 5).value) == 1 or initial) and not flag:
        
        tupdated+=1

        # Extract City and get the values for it
        city = sheet.cell(row = i+1, column = 1).value
        values = get_Value(city)

        # If temperature in celsius is required. 
        if (sheet.cell(row = i+1, column = 4).value) == 'C':

          # Update Humidity
          sheet.cell(row = i+1, column = 3).value =  values[1]
          # Update Temperature
          sheet.cell(row = i+1, column = 2).value =  values[2]
          # Save to file
          wb_obj.save("D:\Programming\Python\Projects\Live Weather Monitor\Values.xlsx") 

          print(f'Location: \t{values[0]}')
          print(f'Humidity: \t{values[1]}')
          print(f'Temperature: \t{values[2]}°C')
          print("----------------------------------------------------")
        
        else:
          # Update Humidity
          sheet.cell(row = i+1, column = 3).value =  values[1]
          # Update Temperature
          sheet.cell(row = i+1, column = 2).value =  toFah(values[2])
          # Save to file
          wb_obj.save("D:\Programming\Python\Projects\Live Weather Monitor\Values.xlsx")

          print(f'Location: \t{values[0]}')
          print(f'Humidity: \t{values[1]}')
          print(f'Temperature: \t{toFah(values[2])}°F')
          print("----------------------------------------------------")

      time.sleep(1)
    initial = False

    if tupdated == 0:
      print("Nothing to update. Exiting!!")
      break
    


if __name__ == "__main__":
    main()
